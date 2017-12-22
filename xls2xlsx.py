#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys
import xlrd
import os.path
from openpyxl.workbook import Workbook as openpyxlWorkbook

def main():
    if len(sys.argv) == 1:
        print('Usage: xls2xlsx.py from.xls')
        return

    bname, ext = os.path.splitext(sys.argv[1])
    if ext != '.xls':
        print('Usage: xls2xlsx.py from.xls')
        return

    xlsBook = xlrd.open_workbook(sys.argv[1])
    workbook = openpyxlWorkbook()

    for i in range(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet = workbook.active if i == 0 else workbook.create_sheet()
        sheet.title = xlsSheet.name

        for row in range(0, xlsSheet.nrows):
            for col in range(0, xlsSheet.ncols):
                sheet.cell(row=row + 1, column=col + 1).value = xlsSheet.cell_value(row, col)
    workbook.save(sys.argv[1] + "x")

if __name__ == '__main__':

     main()
